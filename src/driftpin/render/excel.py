"""Excel renderer: the traceability matrix ships as a first-class sheet,
not a footnote to the test cases.

Requirement IDs are shown using the simple `Req-N` display labels built by
`render/labels.py`, not the registry's real content-addressed IDs.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from driftpin.agents.orchestrator import PipelineResult
from driftpin.render.headers import ArtifactHeader
from driftpin.render.labels import (
    build_requirement_labels,
    label_for,
    labels_for,
    substitute_labels_in_text,
)


def _write_header_sheet(ws: Worksheet, header: ArtifactHeader) -> None:
    ws.title = "Header"
    ws.append(["Field", "Value"])
    ws.append(["Generator", header.generator])
    ws.append(["Generator version", header.generator_version])
    ws.append(["Run ID", header.run_id])
    ws.append(["Registry version", header.registry_version])
    ws.append(["Source document(s)", ", ".join(header.source_doc_titles)])
    ws.append(["Source document hash (for verification)", ", ".join(header.source_doc_hashes)])


def _write_generation_failures_sheet(ws: Worksheet, result: PipelineResult, labels: dict[str, str]) -> None:
    """Only created when at least one scenario's fill call came back empty
    after retries — a `GENERATION_FAILED` scenario needs to read as a
    distinct, attention-grabbing failure, not blend into an ordinary
    zero-coverage row on the Traceability Matrix sheet."""
    scenarios_by_id = {s.scenario_id: s for s in result.strategy.scenarios}
    ws.append(["Scenario ID", "Title", "Status"])
    for scenario_id in result.failed_scenario_ids:
        scenario = scenarios_by_id.get(scenario_id)
        title = substitute_labels_in_text(scenario.title, labels) if scenario else "(unknown scenario)"
        ws.append([scenario_id, title, "GENERATION_FAILED — human attention required"])


def _write_traceability_sheet(ws: Worksheet, result: PipelineResult, labels: dict[str, str]) -> None:
    ws.append(["Requirement", "Title", "Risk Tier", "Coverage Count", "Case IDs"])
    for row in result.traceability:
        ws.append(
            [
                label_for(row.requirement_id, labels),
                row.requirement_title,
                row.risk_tier,
                row.coverage_count,
                ", ".join(row.case_ids),
            ]
        )


def _write_scenarios_sheet(ws: Worksheet, result: PipelineResult, labels: dict[str, str]) -> None:
    ws.append(
        [
            "Scenario ID",
            "Title",
            "Requirements",
            "Owning Agent",
            "Risk Tier",
            "Execution Recommendation",
            "Justification",
        ]
    )
    for scenario in result.strategy.scenarios:
        ws.append(
            [
                scenario.scenario_id,
                substitute_labels_in_text(scenario.title, labels),
                ", ".join(labels_for(scenario.requirement_ids, labels)),
                scenario.owning_agent.value,
                scenario.risk_tier.value,
                scenario.execution_recommendation.value,
                substitute_labels_in_text(scenario.recommendation_justification, labels),
            ]
        )


def _write_test_cases_sheet(ws: Worksheet, result: PipelineResult, labels: dict[str, str]) -> None:
    ws.append(
        [
            "Case ID",
            "Scenario ID",
            "Requirements",
            "Title",
            "Preconditions",
            "Steps",
            "Owning Agent",
            "Execution Recommendation",
            "Assumptions",
        ]
    )
    for case in result.suite.cases:
        steps_text = "; ".join(
            f"{s.step_number}. {substitute_labels_in_text(s.action, labels)} -> "
            f"{substitute_labels_in_text(s.expected_result, labels)}"
            for s in case.steps
        )
        assumptions_text = "; ".join(substitute_labels_in_text(a, labels) for a in case.assumptions)
        ws.append(
            [
                case.case_id,
                case.scenario_id,
                ", ".join(labels_for(case.requirement_ids, labels)),
                substitute_labels_in_text(case.title, labels),
                substitute_labels_in_text(case.preconditions, labels),
                steps_text,
                case.owning_agent.value,
                case.execution_recommendation.value,
                assumptions_text,
            ]
        )


def _write_review_sheet(ws: Worksheet, result: PipelineResult, labels: dict[str, str]) -> None:
    ws.append(["Passed", result.review.passed])
    ws.append(["Summary", substitute_labels_in_text(result.review.summary, labels)])
    ws.append([])
    ws.append(["Severity", "Subject ID", "Description", "Requirements", "Requirement Quote"])
    for finding in result.review.findings:
        ws.append(
            [
                finding.severity.value,
                finding.subject_id,
                substitute_labels_in_text(finding.description, labels),
                ", ".join(labels_for(finding.requirement_ids, labels)),
                substitute_labels_in_text(finding.requirement_quote, labels),
            ]
        )


def build_excel_workbook(result: PipelineResult, header: ArtifactHeader) -> Workbook:
    labels = build_requirement_labels([row.requirement_id for row in result.traceability])

    workbook = Workbook()
    header_sheet = workbook.active
    assert header_sheet is not None
    _write_header_sheet(header_sheet, header)

    if result.failed_scenario_ids:
        _write_generation_failures_sheet(workbook.create_sheet("Generation Failures"), result, labels)

    _write_traceability_sheet(workbook.create_sheet("Traceability Matrix"), result, labels)
    _write_scenarios_sheet(workbook.create_sheet("Scenarios"), result, labels)
    _write_test_cases_sheet(workbook.create_sheet("Test Cases"), result, labels)
    _write_review_sheet(workbook.create_sheet("Review"), result, labels)

    return workbook


def save_excel_workbook(result: PipelineResult, header: ArtifactHeader, output_path: Path) -> None:
    workbook = build_excel_workbook(result, header)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))
