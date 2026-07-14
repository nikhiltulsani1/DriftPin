from __future__ import annotations

from pathlib import Path

import openpyxl

from driftpin.agents.orchestrator import PipelineResult
from driftpin.render.excel import save_excel_workbook
from driftpin.render.headers import build_header
from driftpin.render.markdown import render_markdown_report, render_strategy_markdown
from driftpin.schemas.requirements import Requirement, RiskTier
from driftpin.schemas.review import FindingSeverity, ReviewerFinding, ReviewReport
from driftpin.schemas.strategy import ExecutionRecommendation, OwningAgent, Scenario
from driftpin.schemas.strategy import TestStrategy as StrategyModel
from driftpin.schemas.test_cases import TestCase as CaseModel
from driftpin.schemas.test_cases import TestStep as StepModel
from driftpin.schemas.test_cases import TestSuite as SuiteModel
from driftpin.schemas.test_cases import TraceabilityRow


def _requirement() -> Requirement:
    return Requirement(
        requirement_id="R-abc12345",
        title="Password reset",
        description="Users can reset their password via email.",
        source_span="Users must be able to reset their password via email.",
        source_doc_path="prd.md",
        source_doc_hash="hash-a",
        risk_tier=RiskTier.HIGH,
    )


def _pipeline_result() -> PipelineResult:
    scenario = Scenario(
        scenario_id="S-1",
        title="Password reset flow",
        requirement_ids=["R-abc12345"],
        owning_agent=OwningAgent.FUNCTIONAL_TESTER,
        risk_tier=RiskTier.HIGH,
        execution_recommendation=ExecutionRecommendation.MANUAL,
        recommendation_justification="Low run frequency, high setup cost.",
    )
    case = CaseModel(
        case_id="TC-1",
        scenario_id="S-1",
        requirement_ids=["R-abc12345"],
        title="Reset via emailed link",
        steps=[StepModel(step_number=1, action="Click reset link", expected_result="Reset form shown")],
        owning_agent=OwningAgent.FUNCTIONAL_TESTER,
        execution_recommendation=ExecutionRecommendation.MANUAL,
    )
    return PipelineResult(
        strategy=StrategyModel(strategy_id="strategy-1", scenarios=[scenario]),
        suite=SuiteModel(suite_id="suite-1", strategy_id="strategy-1", cases=[case]),
        review=ReviewReport(
            review_id="review-1",
            target_run_id="run-1",
            findings=[
                ReviewerFinding(
                    severity=FindingSeverity.MINOR,
                    subject_id="S-1",
                    description="Consider a hybrid recommendation.",
                    requirement_ids=["R-abc12345"],
                )
            ],
            passed=True,
            summary="No blockers found.",
        ),
        traceability=[
            TraceabilityRow(
                requirement_id="R-abc12345",
                requirement_title="Password reset",
                risk_tier="high",
                case_ids=["TC-1"],
                coverage_count=1,
            )
        ],
    )


def test_build_header_dedupes_and_sorts_doc_hashes() -> None:
    requirements = [_requirement(), _requirement().model_copy(update={"source_doc_hash": "hash-b"})]
    header = build_header(run_id="run-1", requirements=requirements, registry_version=3)

    assert header.run_id == "run-1"
    assert header.registry_version == 3
    assert header.source_doc_hashes == ["hash-a", "hash-b"]


def test_build_header_includes_source_doc_titles() -> None:
    requirements = [_requirement()]
    header = build_header(run_id="run-1", requirements=requirements, registry_version=1)

    assert header.source_doc_titles == ["prd.md"]


def test_render_markdown_report_contains_all_sections() -> None:
    header = build_header(run_id="run-1", requirements=[_requirement()], registry_version=1)
    report = render_markdown_report(_pipeline_result(), header)

    assert "# Driftpin Test Report" in report
    assert "## Traceability Matrix" in report
    assert "Req-1" in report
    assert "R-abc12345" not in report
    assert "## Scenarios" in report
    assert "S-1" in report
    assert "## Test Cases" in report
    assert "TC-1" in report
    assert "## Review" in report
    assert "No blockers found." in report


def test_render_markdown_report_substitutes_ids_in_review_summary_prose() -> None:
    header = build_header(run_id="run-1", requirements=[_requirement()], registry_version=1)
    result = _pipeline_result()
    result = result.model_copy(
        update={
            "review": result.review.model_copy(
                update={"summary": "Coverage gap on R-abc12345 (Password reset)."}
            )
        }
    )

    report = render_markdown_report(result, header)

    assert "Coverage gap on Req-1 (Password reset)." in report
    assert "R-abc12345" not in report


def test_render_strategy_markdown_contains_header_and_scenarios_only() -> None:
    header = build_header(run_id="run-1", requirements=[_requirement()], registry_version=1)
    strategy = _pipeline_result().strategy

    report = render_strategy_markdown(strategy, header)

    assert "# Driftpin Test Report" in report
    assert "Source document(s): prd.md" in report
    assert "## Scenarios" in report
    assert "S-1" in report
    assert "Req-1" in report
    assert "R-abc12345" not in report
    assert "## Traceability Matrix" not in report
    assert "## Test Cases" not in report
    assert "## Review" not in report


def test_render_markdown_report_shows_generation_failures_when_present() -> None:
    header = build_header(run_id="run-1", requirements=[_requirement()], registry_version=1)
    result = _pipeline_result()
    result = result.model_copy(update={"failed_scenario_ids": ["S-1"]})

    report = render_markdown_report(result, header)

    assert "## Generation Failures" in report
    assert "GENERATION_FAILED" in report
    assert "S-1" in report.split("## Generation Failures")[1].split("## Traceability Matrix")[0]


def test_render_markdown_report_omits_generation_failures_section_when_absent() -> None:
    header = build_header(run_id="run-1", requirements=[_requirement()], registry_version=1)

    report = render_markdown_report(_pipeline_result(), header)

    assert "## Generation Failures" not in report


def test_render_markdown_report_shows_case_assumptions_when_present() -> None:
    header = build_header(run_id="run-1", requirements=[_requirement()], registry_version=1)
    result = _pipeline_result()
    case_with_assumption = result.suite.cases[0].model_copy(
        update={"assumptions": ["ASSUMED: rate limit is 5 attempts — not specified in R-abc12345"]}
    )
    result = result.model_copy(update={"suite": result.suite.model_copy(update={"cases": [case_with_assumption]})})

    report = render_markdown_report(result, header)

    assert "Assumptions:" in report
    assert "rate limit is 5 attempts" in report


def test_render_markdown_report_shows_requirement_quote_on_findings() -> None:
    header = build_header(run_id="run-1", requirements=[_requirement()], registry_version=1)
    result = _pipeline_result()
    finding_with_quote = result.review.findings[0].model_copy(
        update={"requirement_quote": "never silently dropped"}
    )
    result = result.model_copy(update={"review": result.review.model_copy(update={"findings": [finding_with_quote]})})

    report = render_markdown_report(result, header)

    assert 'quote: "never silently dropped"' in report


def test_save_excel_workbook_writes_expected_sheets(tmp_path: Path) -> None:
    header = build_header(run_id="run-1", requirements=[_requirement()], registry_version=1)
    output_path = tmp_path / "report.xlsx"

    save_excel_workbook(_pipeline_result(), header, output_path)

    assert output_path.exists()
    workbook = openpyxl.load_workbook(str(output_path))
    assert workbook.sheetnames == ["Header", "Traceability Matrix", "Scenarios", "Test Cases", "Review"]

    trace_sheet = workbook["Traceability Matrix"]
    assert trace_sheet["A1"].value == "Requirement"
    assert trace_sheet["A2"].value == "Req-1"

    cases_sheet = workbook["Test Cases"]
    assert cases_sheet["A2"].value == "TC-1"


def test_save_excel_workbook_adds_generation_failures_sheet_when_present(tmp_path: Path) -> None:
    header = build_header(run_id="run-1", requirements=[_requirement()], registry_version=1)
    output_path = tmp_path / "report.xlsx"
    result = _pipeline_result().model_copy(update={"failed_scenario_ids": ["S-1"]})

    save_excel_workbook(result, header, output_path)

    workbook = openpyxl.load_workbook(str(output_path))
    assert "Generation Failures" in workbook.sheetnames

    failures_sheet = workbook["Generation Failures"]
    assert failures_sheet["A2"].value == "S-1"
    assert failures_sheet["C2"].value == "GENERATION_FAILED — human attention required"


def test_save_excel_workbook_writes_case_assumptions_column(tmp_path: Path) -> None:
    header = build_header(run_id="run-1", requirements=[_requirement()], registry_version=1)
    output_path = tmp_path / "report.xlsx"
    result = _pipeline_result()
    case_with_assumption = result.suite.cases[0].model_copy(
        update={"assumptions": ["ASSUMED: rate limit is 5 attempts — not specified in R-abc12345"]}
    )
    result = result.model_copy(update={"suite": result.suite.model_copy(update={"cases": [case_with_assumption]})})

    save_excel_workbook(result, header, output_path)

    workbook = openpyxl.load_workbook(str(output_path))
    cases_sheet = workbook["Test Cases"]
    assert cases_sheet["I1"].value == "Assumptions"
    assert "rate limit is 5 attempts" in cases_sheet["I2"].value


def test_save_excel_workbook_writes_requirement_quote_column(tmp_path: Path) -> None:
    header = build_header(run_id="run-1", requirements=[_requirement()], registry_version=1)
    output_path = tmp_path / "report.xlsx"
    result = _pipeline_result()
    finding_with_quote = result.review.findings[0].model_copy(
        update={"requirement_quote": "never silently dropped"}
    )
    result = result.model_copy(update={"review": result.review.model_copy(update={"findings": [finding_with_quote]})})

    save_excel_workbook(result, header, output_path)

    workbook = openpyxl.load_workbook(str(output_path))
    review_sheet = workbook["Review"]
    assert review_sheet["E4"].value == "Requirement Quote"
    assert review_sheet["E5"].value == "never silently dropped"
